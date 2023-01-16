#importing our library
import openpyxl
from openpyxl import Workbook  
from openpyxl.styles import Alignment, Color, colors, PatternFill, Font, Border
from openpyxl.cell import cell
from openpyxl.styles import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension
try:
        from openpyxl.cell import get_column_letter
except ImportError:
        from openpyxl.utils import get_column_letter

your_workbook = Workbook()    #creating the workbook
sheet = your_workbook.active
sheet.merge_cells('A1:B1')
sheet["A1"] = "Col_Name"
sheet["A2"] = "Col_Name"
sheet["A3"] = "Col_Name"
sheet["A4"] = "Col_Name"
sheet["A5"] = "Col_Name"
sheet["A6"] = "Col_Names"
sheet["A7"] = "Col_Name"
sheet["A8"] = "Col_Name"
sheet["A9"] = "Col_Name"
sheet["A10"] = "Col_Name"
sheet["A11"] = "Col_Name"
sheet["A12"] = "Col_Name"
sheet["A13"] = "Col_Name"
sheet["A15"] = "Col_Name"
sheet["B15"] = "Col_Name"
sheet["A17"] = ""
sheet["B17"] = "Col_Name"
sheet["C17"] = "Col_Name"
sheet["D17"] = "Col_Name" 
sheet["E17"] = "Col_Name"
sheet["A18"] = "Col_Name"
sheet["A19"] = "Col_Name"
sheet["A21"] = "Col_Name"
sheet["B21"] = "Col_Name"
sheet["A23"] = ""
sheet["B23"] = "Col_Name"
sheet["C23"] = "Col_Name"
sheet["D23"] = "Col_Name"
sheet["E23"] = "Col_Name"
sheet["A24"] = "Col_Name"
sheet["A25"] = "Col_Name"
sheet["A27"] = "Col_Name"
sheet["B27"] = "Col_Name"
sheet["A29"] = ""
sheet["B29"] = "Col_Name"
sheet["C29"] = "Col_Name"
sheet["D29"] = "Col_Name" 
sheet["E29"] = "Col_Name"
sheet["A30"] = "Col_Name"
sheet["A31"] = "Col_Name"
sheet["A33"] = "Col_Name"
sheet["A35"] = "Statement."
sheet["A36"] = "Statement 2"

#Fill the cell with color
color_fill = PatternFill(start_color='FFFA2A', fill_type='solid')
sheet['A1'].fill = color_fill #assign the column want to fill with color
sheet["A1"].alignment = Alignment(horizontal='center') #cell to be centered

double = Side(border_style="double", color="000000")
thin = Side(border_style="thin", color="000000")
regular = Side(border_style="medium", color="000000")

## For the title cells B2 to F2
for c in sheet['A1:B1'][0]:
    c.border = Border(bottom=double, top=double)

no_left_side = Border(top = regular,bottom=regular,right=regular)
no_right_side = Border(top = regular,bottom=regular, left=regular)
box = Border(top = regular,bottom=regular, left=regular,right=regular)

## For the "table-like" cells
for c in sheet['A2:A13']+sheet['B2:B13']:
    c[0].border = no_left_side
    
for c in sheet['A17:A19']+sheet['B17:B19']+sheet['C17:C19']+sheet['D17:D19']+sheet['E17:E19']:
    c[0].border = no_left_side

for c in sheet['A23:A25']+sheet['B23:B25']+sheet['C23:C25']+sheet['D23:D25']+sheet['E23:E25']:
    c[0].border = no_left_side

for c in sheet['A29:A31']+sheet['B29:B31']+sheet['C29:C31']+sheet['D29:D31']+sheet['E29:E31']:
    c[0].border = no_left_side
    
    
sheet['A35'].alignment = Alignment(wrapText=True)
sheet['A35'].alignment = Alignment(horizontal='left', vertical ='center')
sheet['A1'].font = Font(bold=True)

for column_cells in sheet.columns:
    new_column_length = max(len(str(cell.value)) for cell in column_cells)
    new_column_letter = (get_column_letter(column_cells[0].column))
    if new_column_length > 0:
        sheet.column_dimensions[new_column_letter].width = new_column_length*0.5
        
        
for b in sheet['A2:A33']+sheet['B2:B33']:
    b[0].font = Font(bold=True)
    
for b in sheet['A17:A19']+sheet['B17:B19']+sheet['C17:C19']+sheet['D17:D19']+sheet['E17:E19']:
    b[0].font = Font(bold=True)

for b in sheet['A23:A25']+sheet['B23:B25']+sheet['C23:C25']+sheet['D23:D25']+sheet['E23:E25']:
    b[0].font = Font(bold=True)

for b in sheet['A29:A31']+sheet['B29:B31']+sheet['C29:C31']+sheet['D29:D31']+sheet['E29:E31']:
    b[0].font = Font(bold=True)
    

file = your_workbook.save("C:\\Users\\File_path\\file.xlsx") #saving the file with the 'xlsx' excel extension
